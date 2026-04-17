from typing import Annotated, Optional
from io import BytesIO

from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook, load_workbook

from app.database import get_db
from app.dependencies import get_current_user, require_role
from app.models.user import User
from app.schemas.contact import ContactCreate, ContactUpdate, ArchiveRequest
from app.schemas.activity import ActivityCreate
from app.services import contact_service, activity_service
from app.utils.response import ok, fail

router = APIRouter()


@router.get("")
async def list_contacts(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    search: Optional[str] = Query(None),
    industry: Optional[str] = Query(None),
    assigned_to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=500),
    sort_by: str = Query("created_at"),
    order: str = Query("desc"),
    is_archived: int = Query(0),
):
    result = await contact_service.list_contacts(
        db, current_user, search, industry,
        assigned_to, page, page_size, sort_by, order, is_archived,
    )
    return ok(data=result)


@router.get("/import/template")
async def download_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "customer template"
    headers = [
        "name", "company", "industry", "email", "phone", "address", "remark",
        "tag", "assigned_to_email",
        "status", "priority", "amount", "deal_title",
        "customer_id (add deal only)",
    ]
    ws.append(headers)
    # Example row — new contact
    ws.append([
        "Zhang San", "Example Tech Co.", "Technology/IT",
        "zhangsan@example.com", "13800138000", "123 Main St, Beijing",
        "Example note", "VIP,key account", "sales@crm.com",
        "lead", "mid", "100000", "", "",
    ])
    # Example row — add deal to existing contact
    ws.append([
        "", "", "", "", "", "", "", "", "",
        "negotiating", "high", "50000", "Q2 renewal", "existing-contact-uuid",
    ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="contact_import_template.xlsx"'},
    )


@router.post("/import")
async def import_contacts(
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
    file: UploadFile = File(...),
):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return fail("Only .xlsx or .xls files are supported", status_code=400)

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        return fail("File size cannot exceed 10MB", status_code=400)

    wb = load_workbook(BytesIO(content), read_only=True)
    ws = wb.active

    field_map = {
        "name": "name", "company": "company", "industry": "industry",
        "email": "email", "phone": "phone", "address": "address", "remark": "notes",
        "tag": "tags", "assigned_to_email": "assigned_to_email",
        "status": "status", "priority": "priority", "amount": "deal_value",
        "deal_title": "deal_title",
        "customer_id (add deal only)": "id",
    }

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return fail("File is empty", status_code=400)

    # Map headers
    col_map = {}
    for idx, cell in enumerate(header):
        cell_str = str(cell or "").strip()
        if cell_str in field_map:
            col_map[idx] = field_map[cell_str]

    parsed_rows = []
    for row in rows_iter:
        row_dict = {}
        for idx, field_name in col_map.items():
            val = row[idx] if idx < len(row) else None
            row_dict[field_name] = str(val).strip() if val is not None else ""
        if any(v for v in row_dict.values()):
            parsed_rows.append(row_dict)

    wb.close()

    result = await contact_service.import_contacts(db, parsed_rows, current_user)
    return ok(data=result)


@router.get("/{contact_id}")
async def get_contact(
    contact_id: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _current_user: Annotated[User, Depends(get_current_user)],
):
    contact = await contact_service.get_contact(db, contact_id)
    if not contact:
        return fail("Contact not found", code="NOT_FOUND", status_code=404)
    return ok(data=contact)


@router.post("")
async def create_contact(
    body: ContactCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    data = body.model_dump()
    contact = await contact_service.create_contact(db, data, current_user)
    return ok(data=contact, message="Contact created", status_code=201)


@router.put("/{contact_id}")
async def update_contact(
    contact_id: str,
    body: ContactUpdate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    data = {k: v for k, v in body.model_dump().items() if v is not None}
    if "is_archived" in data and current_user.role not in ("admin", "manager"):
        return fail("Permission denied", code="FORBIDDEN", status_code=403)
    contact = await contact_service.update_contact(db, contact_id, data, current_user)
    if not contact:
        return fail("Contact not found", code="NOT_FOUND", status_code=404)
    return ok(data=contact, message="Updated successfully")


@router.patch("/{contact_id}/archive")
async def archive_contact(
    contact_id: str,
    body: ArchiveRequest,
    db: Annotated[AsyncSession, Depends(get_db)],
    _current_user: Annotated[User, Depends(require_role("admin", "manager"))],
):
    contact = await contact_service.archive_contact(db, contact_id, body.is_archived)
    if not contact:
        return fail("Contact not found", code="NOT_FOUND", status_code=404)
    msg = "Archived" if body.is_archived else "Unarchived"
    return ok(data=contact, message=f"{msg} successfully")


@router.delete("/{contact_id}")
async def delete_contact(
    contact_id: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _current_user: Annotated[User, Depends(require_role("admin"))],
):
    deleted = await contact_service.soft_delete_contact(db, contact_id)
    if not deleted:
        return fail("Contact not found", code="NOT_FOUND", status_code=404)
    return ok(message="Deleted successfully")


# --- Activity routes ---

@router.get("/{contact_id}/activities")
async def list_activities(
    contact_id: str,
    db: Annotated[AsyncSession, Depends(get_db)],
    _current_user: Annotated[User, Depends(get_current_user)],
):
    activities = await activity_service.list_activities(db, contact_id)
    return ok(data=activities)


@router.post("/{contact_id}/activities")
async def create_activity(
    contact_id: str,
    body: ActivityCreate,
    db: Annotated[AsyncSession, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_user)],
):
    activity = await activity_service.create_activity(
        db, contact_id, body.deal_id, current_user.id, body.type, body.content, body.follow_date,
    )
    return ok(data=activity, message="Follow-up activity recorded", status_code=201)
